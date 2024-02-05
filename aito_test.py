import openpyxl

wb = openpyxl.load_workbook('sample.xlsx')
ws = wb['Sheet']

i = 1

for row in ws.iter_rows():
    for cell in row:
        if cell.row == 1:
            ws.cell(row=31, column=i).value = cell.value
            i = i + 1

print(cell.value)

#別名で保存
wb.save('sample.xlsx')