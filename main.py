import qrcode
import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import openpyxl as op
#import pandas as pd
import datetime
import numpy as np
import pyautogui as pa
from playsound import playsound
from openpyxl.styles import PatternFill
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import os
# import PIL
import random

column = ["A", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
       "AA", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", 
       "BA", "BC", "BD", "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU", "BV", "BW", "BX", "BY", "BZ"]

row = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', 
          '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', 
          '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', 
          '31', '32', '33', '34', '35', '36', '37', '38', '39', '40', 
          '41', '42', '43', '44', '45', '46', '47', '48', '49', '50', 
          '51', '52', '53', '54', '55', '56', '57', '58', '59', '60', 
          '61', '62', '63', '64', '65', '66', '67', '68', '69', '70', 
          '71', '72', '73', '74', '75', '76', '77', '78', '79', '80', 
          '81', '82', '83', '84', '85', '86', '87', '88', '89', '90', 
          '91', '92', '93', '94', '95', '96', '97', '98', '99', '100']




# name = [
#      '   ' , '明石凛太郎', '伊藤匠真', '伊藤千晴' , '今村 泰志' , '入江 麗姫' , 
#      '浦塚 武士' , '江頭 弦樹' , '大坪 諒一' , '梶原 徳真' , '加藤田 悠真'  , 
#      '河津 仁晴' , '久保田 勇気' , '江田 光来' , '古賀 結莉華' , '後藤 博喜' , 
#      '小林 士秩' , '是松 文太郎' , '坂井 瑠衣' , '貞方 海生' , '佐野 鈴莉' , 
#      '猿渡 希流' , '白水 希龍' , '末續 藍人' , '田中 健翔' , '田中 志保里' , 
#      '谷口 輝太朗' , '寺島 大虎' , '寺元 一耕' , '土井 秀哉' , '富田 涼介' , 
#      '内藤 誉起' , '中村 開斗' , '西田 大和人' , '西村 雅基' , '花岡 瑞生' , 
#      '日吉 曜大' , '平川 粹' , '藤井 颯' , '松永 光世' , '松吉 宇宙' , 
#      '三好 俊祐' , '森山 優星' , '山下 紘輝' , '吉山 響' , '渡辺 嵩広' , 
#      'アマラ'
#      ]

st.title('QR出席管理')
number_list = []

with open("./name_data.txt",mode="r",encoding="utf-8") as f:
    lines = f.readlines()
    name = []
    for word in lines:
        name.append(word.replace("\n",""))


if st.checkbox('名前を追加する'):
    with st.form("myform",clear_on_submit=True):
        newname = st.text_input('名前を追加する')
        if st.form_submit_button("追加"):
            name = list(filter(None,name))
            name.append(newname)
            # x, y = pa.locateCenterOnScreen('button2.png')
            # pa.leftClick(x, y)
            pa.press("F5")


with open("./name_data.txt",mode="w",encoding="utf-8") as f:
    namerabel = ""
    for word in name:
        namerabel = namerabel + word + "\n"
    f.write(namerabel)

namelength = len(name)
# print(name)



###エクセル作成など


namelist = op.Workbook()
book = op.load_workbook('C:\STREAMLIT\streamlit-sample\sample.xlsx')
project = book.worksheets[0]
project.cell(row = 1, column = 1, value = "出席番号")
project.cell(row = 1, column = 2, value = "名前")


# if st.checkbox('新しいシートを作成します'):
#     book.create_sheet('Sheet')

# i = 1
# kadai = [" "]
# for row in project.iter_rows(min_col = 3, max_row = 1):
#     for col in row:
#         kadai.append(col.value)

master_kadai = []
with open("./exam_name.txt",mode="r",encoding="utf-8") as g:
    lines = g.readlines()
    exam_title = []
    for word in lines:
        exam_title.append(word.replace("\n", ""))

newkadai = list(filter(None, exam_title))
kari = (list(filter(None, exam_title)))
master_kadai.append(kari)
# newkadai.replace("\n", "")

kadai_volume = len(newkadai)
# print(kari)
# print(newkadai)














with open("./permanent_exam_name.txt", mode = "r", encoding = "utf-8") as k:
    lines = k.readlines()
    permanent_exam_name = []
    master_note = []
    for word in lines:
        permanent_exam_name.append(word.replace("\n", ""))
        master_note.append(word.replace("\n", ""))



if st.checkbox("ノート課題を追加する"):
    with st.form("permanentform", clear_on_submit=True):
        permanent_title = st.text_input("ノート課題を追加する")
        if st.form_submit_button("追加"):
            permanent_exam_name = list(filter(None,permanent_exam_name))
            permanent_exam_name.append(permanent_title)
            master_note.append(word.replace("\n", ""))
            book.create_sheet(index = 1, title = permanent_title)
            # sheet_volume = len(book.worksheets) + 1
            # book.create_sheet(index = sheet_volume, title = permanent_title)
            # pa.press("F5")
            ##課題を新規作成した時のQRコードをファイルに保存(ノート)
            qrcnt = 1
            make_dir_path = "C:\STREAMLIT\streamlit-sample\image\\" + permanent_exam_name[-1]

            if os.path.isdir(make_dir_path):
                    print("note")
            else:
                os.makedirs(make_dir_path)
                qrcnt = 1
                master_note_volume = len(master_note)
                strmaster_note_volume = str(master_note_volume)


                for i in range(namelength):
                    saveqr_data = str(qrcnt) + " " + "0" + " " + strmaster_note_volume
                    saveqr = qrcode.make(data = saveqr_data, version = 1)
                    saveqr.save(make_dir_path + "\\" + str(qrcnt) + ".jpg", label = str(qrcnt), quality = 75, optimize = True, progressive = True)
                    
                    img = Image.open(make_dir_path + "\\" + str(qrcnt) + ".jpg")
                    draw = ImageDraw.Draw(img)
                    font = ImageFont.truetype("C:\\meiryo.ttc", 30)
                    draw.text((150, 300), str(qrcnt) + permanent_exam_name[-1], anchor = 'md', font = font)
                    img.save(make_dir_path + "\\" + str(qrcnt) + ".jpg")
                    # saveqr.show()
                        
                    qrcnt = qrcnt + 1


with open("./permanent_exam_name.txt",mode="w",encoding="utf-8") as k:
    examrabel = ""
    for word in permanent_exam_name:
        examrabel = examrabel + word + "\n"
    k.write(examrabel)
    



permanent_cnt = 1
permanent_volume = len(permanent_exam_name)

def startwrite():
    for idx in range(1, namelength):
        Z = name[idx]

        note.cell(row=idx+2, column=1, value=idx)
        note.cell(row=idx+2, column=2, value=Z)

    for i in range(0, kadai_volume):
        F = newkadai[i]
        project.cell(row=1, column=i + 3, value=F)
    


for i in range(permanent_volume):

    book.create_sheet(index = 1, title = permanent_exam_name[permanent_cnt - 1])
    sheetname = permanent_exam_name[permanent_cnt - 1]
    note = book[sheetname]
    note.cell(column = 1, row = 1, value = sheetname)
    note.cell(row = 2, column = 1, value = "出席番号")
    note.cell(row = 2, column = 2, value = "名前")
    startwrite()


    # sheetname = sheetname + "1"

    # book.remove(sheetname)
    # if book != "Sheet1" or sheetname:
    #     book.remove(book)
    # for ws in book.worksheets:
    #     if ws.title != "Sheet1" or sheetname:
    #         book.remove(index = permanent_cnt + 1)
    #         book.save('sample.xlsx')

    # # book.remove(book.worksheets[1])
    # book.save('sample.xlsx')
    

    permanent_cnt = permanent_cnt + 1




# for i in range(30):
#     D = 0
#     ws = book[D]
#     if ws.values == None:
#         ws.defined_names
#     book.save('sample.xlsx')



# D = 0
# for i in range(len(ws_sheetname)):
#     ws = book.worksheets[D]
#     if ws.values == None:
#         book.remove_sheet(ws_sheetname[D])
#     # print(ws.values)
#     D = D + 1
    


ws_sheetname = book.sheetnames

for i in range(len(ws_sheetname) - permanent_volume - 1):
    ws = book.worksheets[1]
    book.remove(ws)

time_now = datetime.datetime.now()
time = str(time_now.month) + "/" + str(time_now.day) + " " + str(time_now.hour) + ":" + str(time_now.minute)


##ノート課題日付
# aa

calender_cnt = 3
calender = book.worksheets[1]
calender_list = []
calender_value = calender.cell(row = 2, column = calender_cnt).value

while calender_value != None:
    calender_value = calender.cell(row = 2, column = calender_cnt).value
    calender_list.append(calender_value)
    calender_cnt = calender_cnt + 1

calender_list = list(filter(None,calender_list))

# print(calender_list)




# aa






###課題を追加する


if st.checkbox('課題を追加する'):
    with st.form("appendform", clear_on_submit=True):
        exam_title = st.text_input('課題を追加する')
        
        project.cell(row=1, column=kadai_volume + 3, value=exam_title)
        submitted = st.form_submit_button("追加")
        if submitted:
            newkadai.append(exam_title)
            master_kadai.append(exam_title)
            ##課題を新規作成した時のQRコードをファイルに保存(課題)

            qrcnt = 1
            make_dir_path = "C:\STREAMLIT\streamlit-sample\image\\" + newkadai[-1]

            if os.path.isdir(make_dir_path):
                    print("nande")
            else:
                print("kotti")
                os.makedirs(make_dir_path)
                qrcnt = 1
                print(kari)
                kari_volume = len(kari) + 1
                print(kari_volume)
                strkari_volume = str(kari_volume)
                print(strkari_volume)
                
                for i in range(namelength):
                    saveqr_data = str(qrcnt) + " " + str(len(newkadai)) + " " + "0"
                    saveqr = qrcode.make(data = saveqr_data, version = 1)
                    saveqr.save(make_dir_path + "\\" + str(qrcnt) + ".jpg", label = str(qrcnt), quality = 75, optimize = True, progressive = True)
                    
                    img = Image.open(make_dir_path + "\\" + str(qrcnt) + ".jpg")
                    draw = ImageDraw.Draw(img)
                    font = ImageFont.truetype("C:\\meiryo.ttc", 30)
                    draw.text((150, 300), str(qrcnt) + newkadai[-1], anchor = 'md', font = font)
                    img.save(make_dir_path + "\\" + str(qrcnt) + ".jpg")
                    # saveqr.show()
                        
                    qrcnt = qrcnt + 1  
            pa.press("F5")
        
    
with open("./exam_name.txt",mode="w",encoding="utf-8") as g:
    examrabel = ""
    for word in newkadai:
        examrabel = examrabel + word + "\n"
    g.write(examrabel)
    

# allkadai = newkadai + permanent_exam_name


student_number = 0
string_number = str(student_number)

###qrコードの生成
a = st.sidebar.write('あなたの出席番号を教えてください')


number = st.sidebar.selectbox(
    'あなたの出席番号を教えてください',
    list(range(0,namelength))
)
# purinto = st.sidebar.toggle("プリント課題を提出する")


display_newkadai = newkadai.insert(0, "選択してください")
qrkadai = st.sidebar.selectbox(
    '課題選択',
    list(newkadai)
)

# del newkadai[0]
# print(display_newkadai)
# print(newkadai)
# print(qrkadai)


# if purinto:
#     selectprinto()


display_notekadai = permanent_exam_name.insert(0, "選択してください")
notekadai = st.sidebar.selectbox(
    "ノート課題選択", 
    list(permanent_exam_name)
)   

qrkadai2 = newkadai.index(qrkadai) - 1      #####################
notekadai2 = permanent_exam_name.index(notekadai) - 1


b = name[number]
# st.title(b)
if number <= 0:
    st.sidebar.write(b)
else:
    st.sidebar.write('出席番号',  number,  '番の', b , 'さんこんにちは！！')
    time_now = datetime.datetime.now()
    st.sidebar.write(time_now)

exam_number = len(newkadai)                                 #########################################

strnumber = str(number)
strqrkadai2 = str(qrkadai2 + 1)
strnotekadai2 = str(notekadai2 + 1)

QRdata = strnumber + " " + strqrkadai2 + " " + strnotekadai2

st.sidebar.write(QRdata)
# print(QRdata)

qr = qrcode.make(data = QRdata , version = 1)
img = Image.open('number.png')
qr.save('number.png')

# draw = ImageDraw.Draw(img)
# font = ImageFont.truetype("C:\\meiryo.ttc", 30)
# draw.text((150, 300), newkadai[qrkadai2 + 1], anchor = 'md', font = font)
# img.save('number.png')

# dd
# print(exam_number)

if st.sidebar.checkbox('QRコードを表示しますか？'):
    if number == 0:
         st.sidebar.write("出席番号を選択してください")
        
    elif notekadai2 == -1 and qrkadai2 == -1:
         st.sidebar.write("課題を選択してください")

    elif notekadai2 > -1 and qrkadai2 > -1:
        st.sidebar.write("課題かノート課題どちらかのみを選択してください")

    else:
        img = Image.open('number.png')
        st.sidebar.image(img, caption = strnumber + '番のQRコード', use_column_width = False)

# print(qrkadai2)


##ファイルにQRコード作成・保存
qrcnt = 1
kadai_idx = 0
folder_path = "C:\STREAMLIT\streamlit-sample\image"


idx = 1
for i in range(len(kari)):
    make_dir_path = "C:\STREAMLIT\streamlit-sample\image\\" + newkadai[idx]   
    idx = idx + 1
    if os.path.isdir(make_dir_path):
        pass
    else:
        os.makedirs(make_dir_path)
        qrcnt = 1
        kadai_idx = kadai_idx + 1
        strkadai_idx = str(kadai_idx)

        for i in range(namelength):
            saveqr_data = str(qrcnt) + " " + strkadai_idx + " " + "0"
            saveqr = qrcode.make(data = saveqr_data, version = 1)
            saveqr.save(make_dir_path + "\\" + str(qrcnt) + ".jpg", label = str(qrcnt), quality = 75, optimize = True, progressive = True)
            img = Image.open(make_dir_path + "\\" + str(qrcnt) + ".jpg")
            draw = ImageDraw.Draw(img)
            font = ImageFont.truetype("C:\\meiryo.ttc", 30)
            draw.text((150, 300), str(qrcnt) + newkadai[idx - 1], anchor = 'md', font = font)
            img.save(make_dir_path + "\\" + str(qrcnt) + ".jpg")
            # saveqr.show()
            
            qrcnt = qrcnt + 1


##ノートのQR作成・保存

sound_list = ["get.mp3", "mario-1up.mp3"]

    # eara
qrcnt = 1
kadai_idx = 0
idx = 1
for i in range(len(master_note)):
    make_dir_path = "C:\STREAMLIT\streamlit-sample\image\\" + permanent_exam_name[idx]   
    idx = idx + 1
    if os.path.isdir(make_dir_path):
        pass
    else:
        os.makedirs(make_dir_path)
        qrcnt = 1
        kadai_idx = kadai_idx + 1
        strkadai_idx = str(kadai_idx)

        for i in range(namelength):
            saveqr_data = str(qrcnt) + " " + "0" + " " + strkadai_idx
            saveqr = qrcode.make(data = saveqr_data, version = 1)
            saveqr.save(make_dir_path + "\\" + str(qrcnt) + ".jpg", label = str(qrcnt), quality = 75, optimize = True, progressive = True)
            
            img = Image.open(make_dir_path + "\\" + str(qrcnt) + ".jpg")
            draw = ImageDraw.Draw(img)
            font = ImageFont.truetype("C:\\meiryo.ttc", 30)
            draw.text((150, 300), str(qrcnt) + permanent_exam_name[idx - 1], anchor = 'md', font = font)
            img.save(make_dir_path + "\\" + str(qrcnt) + ".jpg")
            # saveqr.show()
            
            qrcnt = qrcnt + 1                

##課題を新規作成した時のQRコードをファイルに保存(課題)

# qrcnt = 1
# make_dir_path = "C:\STREAMLIT\streamlit-sample\image\\" + newkadai[-1]

# if os.path.isdir(make_dir_path):
#         print("nande")
# else:
#     print("kotti")
#     os.makedirs(make_dir_path)
#     qrcnt = 1
#     print(kari)
#     kari_volume = len(kari) + 1
#     print(kari_volume)
#     strkari_volume = str(kari_volume)
#     print(strkari_volume)
    
#     for i in range(namelength):
#         saveqr_data = str(qrcnt) + " " + str(len(newkadai)) + " " + "0"
#         saveqr = qrcode.make(data = saveqr_data, version = 1)
#         saveqr.save(make_dir_path + "\\" + str(qrcnt) + ".jpg", label = str(qrcnt), quality = 75, optimize = True, progressive = True)
#         # saveqr.show()
            
#         qrcnt = qrcnt + 1  

# ##課題を新規作成した時のQRコードをファイルに保存(ノート)
# qrcnt = 1
# make_dir_path = "C:\STREAMLIT\streamlit-sample\image\\" + permanent_exam_name[-1]

# if os.path.isdir(make_dir_path):
#         print("note")
# else:
#     os.makedirs(make_dir_path)
#     qrcnt = 1
#     master_note_volume = len(master_note) + 1
#     strmaster_note_volume = str(master_note_volume)


#     for i in range(namelength):
#         saveqr_data = str(qrcnt) + " " + "0" + " " + strmaster_note_volume
#         saveqr = qrcode.make(data = saveqr_data, version = 1)
#         saveqr.save(make_dir_path + "\\" + str(qrcnt) + ".jpg", label = str(qrcnt), quality = 75, optimize = True, progressive = True)
#         # saveqr.show()
            
#         qrcnt = qrcnt + 1  

















# mecchaera-haiterukoko


# if os.path.isdir(make_dir_path):
#     print("koko")
#     pass
# else:
#     os.makedirs(make_dir_path)






# if st.checkbox("QRコードを生成する"):
#     img = Image.open('number.png')
#     st.image(img, caption = strnumber + '番のQRコード', use_column_width = False)








# if __name__ == "__main__":
#     image00 = Image.open('C:\STREAMLIT\streamlit-sample\kote00.jpg')
#     image01 = Image.open('C:\STREAMLIT\streamlit-sample\kote01.jpg')
#     image02 = Image.open('C:\STREAMLIT\streamlit-sample\kote02.jpg')
#     image03 = Image.open('C:\STREAMLIT\streamlit-sample\kote03.jpg')
#     image04 = Image.open('C:\STREAMLIT\streamlit-sample\kote04.jpg')

#     images = [image00, image01, image02, image03, image04]

#     fig = plt.figure()
#     for i, im in enumerate(images):
#         fig.add_subplot(4,5,i+1).set_title(str(i))
#         plt.imshow(im)
    
#     plt.show()




###エクセル初期設定
for idx in range(1, namelength):
    x = name[idx]

    project.cell(row=idx+1, column=1, value=idx)
    project.cell(row=idx+1, column=2, value=x)

for i in range(0, kadai_volume):
    Y = newkadai[i + 1]
    project.cell(row=1, column=i + 3, value=Y)

# for idx in range(1, namelength):
#     Z = name[idx]

#     note.cell(row=idx+2, column=1, value=idx)
#     note.cell(row=idx+2, column=2, value=Z)

# for i in range(0, kadai_volume):
#     F = newkadai[i]
#     project.cell(row=1, column=i + 3, value=F)


###QRコードの受付
#エクセルに〇をつける




# unsubmited_color = PatternFill(patternType='solid', fgColor='FFFF00', bgColor= 'FFFF00')



# kadaicnt = 2
newkadaicnt = 0
kadai_submit = []
if st.checkbox("未提出者を表示する"):
    j = st.selectbox(
        "課題を選択する",
        list(newkadai)
    )

    # if j <= 1:
    #     st.write("全員提出しています!")
    k = newkadai.index(j) + 2
    # print(k)
    kadaicnt = 2
    for i in range(namelength):
        cell = project.cell(row = kadaicnt, column = k)
        kadai_submit.append(cell.value)
        kadaicnt = kadaicnt + 1

    until_sutudent = [i for i, x in enumerate(kadai_submit)if x == None]
    until_sutudent_1 = [i + 1 for i in until_sutudent]

    until_sutudent_volume = len(until_sutudent)
    for i in range(1, until_sutudent_volume):
        C = until_sutudent[newkadaicnt] + 1
        D = name[C]
        st.write(str(C) + "番   " + D)
        newkadaicnt = newkadaicnt + 1
    if len(until_sutudent) < 2:
        st.write("課題を選択してください")

    n = st.selectbox(
    "ノート課題を選択する",
    permanent_exam_name
    )
    m = permanent_exam_name.index(n)
    
    # print(m)
    kadaicnt = 3
    note_select = []
    note_sentaku = book.worksheets[-m]

    cell = note_sentaku.cell(row = 2, column = kadaicnt).value
    while cell != None:
        cell = note_sentaku.cell(row = 2, column = kadaicnt).value
        note_select.append(cell)
        kadaicnt = kadaicnt + 1
        # print(note_select)

    note_select = list(filter(None,note_select))
    
    
    
    p = st.selectbox(
        "日付を選択してください",
        note_select
    )

    note_miteishutu = []
    if m == 0:
        pass
    else:
        r = note_select.index(p)
    
        # print(r)
        kadaicnt = 3
        for i in range(namelength):
            # note_sentaku = book.worksheets[-m]
            cell = note_sentaku.cell(row = kadaicnt, column = 3 + r)##clomn ijiru
            note_miteishutu.append(cell.value)
            kadaicnt = kadaicnt + 1
    

        if r == None:
            st.write("全員提出しています")
    
    # print(note_miteishutu)

    note_until_sutudent = [i for i, x in enumerate(note_miteishutu)if x == None]
    note_until_sutudent_1 = [i + 1 for i in note_until_sutudent]


    note_until_sutudent_volume = len(note_until_sutudent)
    newkadaicnt = 0
    for i in range(1, note_until_sutudent_volume):
        C = note_until_sutudent[newkadaicnt] + 1
        D = name[C]
        st.write(str(C) + "番   " + D)
        newkadaicnt = newkadaicnt + 1
    
    print(until_sutudent)

    if len(until_sutudent) < 2 and len(note_until_sutudent) < 2:
        st.write("全員提出しています")




qrread = []
on = st.toggle("提出モード")
if on:
    with st.form("submitform",clear_on_submit=True):
        submission2 = st.text_input("QRコードをかざしてください")
        # pa.press("F11")
        # pa.moveTo(1, 1)
        # pa.move(850,360,duration=0.1)
        # pa.click()
        if st.form_submit_button("提出"):
            # x, y = pa.locateCenterOnScreen('button2.png', confidence = 0.6)
            pa.click(1000, 450)
            qrread = submission2.split()
            read_row = int(qrread[0]) + 1
            read_column_1 = int(qrread[1]) + 2
            read_column_2 = int(qrread[2])

            if int(qrread[1]) != 0 and int(qrread[2]) != 0:
                st.write("error")
            elif int(qrread[1]) == 0 and int(qrread[2]) == 0:
                st.write("error")
            elif int(qrread[0]) == 0:
                st.write("error")
            elif  int(qrread[1]) > 0:
                if time_now.hour >= 9:
                    iro = column[read_column_1 - 2] + row[read_row - 1]
                    print(iro)
                    print("kokoimunida")
                    project.cell(row=read_row, column=read_column_1, value=time)
                    
                    project[iro].font = op.styles.fonts.Font(color = 'FF0000')
                else:
                    project.cell(row=read_row, column=read_column_1, value=time)
                
            else:
                if time_now.hour <= 9:
                    notecode = read_column_2 - (read_column_2 * 2)
                    select_note = book.worksheets[notecode]

                    calender_cnt = 3
                    calender_list = []
                    calender_value = select_note.cell(row = 2, column = calender_cnt).value

                    while calender_value != None:
                        calender_value = select_note.cell(row = 2, column = calender_cnt).value
                        calender_list.append(calender_value)
                        calender_cnt = calender_cnt + 1

                    calender_list = list(filter(None,calender_list))
                    calender_volume = len(calender_list)
                    calender_time = str(time_now.month) + "." + str(time_now.day)
                    # calender_time_2 = "'" + str(time_now.month) + "." + str(time_now.day)
                    calender_time_2 = str(time_now.month) + "." + str(time_now.day)            ### "'" 問題


                    # calender_time = float(calender_time)

                    

                    # print(calender_time)

                    # calender_list_2 = [1.18, 1.19, 1.20]

                    # print(calender_time)
                    print(calender_list)

                    if calender_time in calender_list:
                        calender_number = calender_list.index(calender_time)
                        select_note.cell(row = read_row + 1, column=calender_number + 3, value=time)
                        print("ok")
                    else:
                        select_note.cell(row = 2, column = calender_volume + 3, value = calender_time_2)
                        
                        select_note.cell(row = read_row + 1, column=calender_volume + 3, value=time)

                        print("だめ")
                    
                else:                    
                    notecode = read_column_2 - (read_column_2 * 2)
                    select_note = book.worksheets[notecode]

                    calender_cnt = 3
                    calender_list = []
                    calender_value = select_note.cell(row = 2, column = calender_cnt).value

                    while calender_value != None:
                        calender_value = select_note.cell(row = 2, column = calender_cnt).value
                        calender_list.append(calender_value)
                        calender_cnt = calender_cnt + 1

                    calender_list = list(filter(None,calender_list))
                    calender_volume = len(calender_list)
                    calender_time = str(time_now.month) + "." + str(time_now.day)
                    # calender_time_2 = "'" + str(time_now.month) + "." + str(time_now.day)
                    calender_time_2 = str(time_now.month) + "." + str(time_now.day)            ### "'" 問題


                    # calender_time = float(calender_time)

                    

                    # print(calender_time)

                    # calender_list_2 = [1.18, 1.19, 1.20]

                    # print(calender_time)
                    print(calender_list)

                    # index = []
                    
                    

                    if calender_time in calender_list:
                        calender_number = calender_list.index(calender_time)
                        select_note.cell(row = read_row + 1, column=calender_number + 3, value=time)
                        iro = column[calender_number + 1] + row[read_row]
                        print(iro)
                        select_note[iro].font = op.styles.fonts.Font(color = 'FF0000')
                        
                        print("ok")
                    else:
                        select_note.cell(row = 2, column = calender_volume + 3, value = calender_time_2)
                        
                        select_note.cell(row = read_row + 1, column=calender_volume + 3, value=time)

                        print("だめ")

        




            sound = random.randint(0, 1)

            playsound(sound_list[sound])
            # print(read_row)
            # print(read_column_1)






###未提出者リスト


# n = 0
# for i in range(len(master_note)):
#     master_kadai.append(master_note[n])
#     n = n + 1
master_kadai.append(master_note)
# print(master_kadai)

# master_kadai.append(newkadai)














private_submit = []
cnt = 3
newcnt = 0






# date_cnt = 0
# count = -1
# # count2 = 0
# date_list = []
# note_submit = []
# note_list =[]
# for i in range(permanent_volume):
#     print(count)
#     permanent_worksheet = book.worksheets[count]
#     count = count - 1
#     date_info = permanent_worksheet.cell(row = 2, column = date_cnt + 3).value
#     date_cnt = 0
#     while date_info != None:
#         date_info = permanent_worksheet.cell(row = 2, column = date_cnt + 3).value
#         note_list.append(date_info)
#         date_cnt = date_cnt + 1
#         print(note_list)























if st.sidebar.checkbox(b + "さんの未提出課題を表示しますか？"):
    for i in range(kadai_volume):
        cell = project.cell(row = number + 1, column = cnt)
        private_submit.append(cell.value)
        cnt = cnt + 1
    



    until_submit = [i for i, x in enumerate(private_submit)if x == None]
    until_submit_1 = [i + 1 for i in until_submit]



    date_cnt = 0
    count = -1
    # # count2 = 0
    date_list = []
    note_submit = []
    # note_submit_7 = []
    note_list = []
    note_list_x = []
    note_list_y = []



    for i in range(permanent_volume):
        
        permanent_worksheet = book.worksheets[count]
        count = count - 1
        
        # print(count)
        date_cnt = 0
        date_info = permanent_worksheet.cell(row = 2, column = date_cnt + 3).value
        
        note_cnt = 0
        while date_info != None:
            
            date_info = permanent_worksheet.cell(row = 2, column = date_cnt + 3).value
            date_list.append(date_info)
            date_cnt = date_cnt + 1
            note_list_x.append(date_info)
            
            note_list_x = list(filter(None,note_list_x))
            note_list_y.append(permanent_worksheet["A1"].value)
            note_cnt = note_cnt + 1

            if note_list_x[-1] == str(permanent_worksheet["A1"].value) + str(None):
                del note_list_x[-1]
                note_cnt = note_cnt - 1
            
        del note_list_y[-1]

        note_cnt = note_cnt - 1
            # print(date_list)
        

        
        date_list = list(filter(None,date_list))
        date_volume = len(date_list)
        
        # print(note_list)
        # print(date_volume)
        count2 = 0
        # print(note_cnt)
        for i in range(note_cnt):
            
            note_submit_info = permanent_worksheet.cell(row = number + 2, column = count2 + 3).value
            # if note_submit_info == None:
            note_submit.append(note_submit_info)
            count2 = count2 + 1
            # print(note_submit)
            
        
    
    

    # count = count + 1
            
    note_list = [a + b for (a, b) in zip(note_list_y, note_list_x)]
    # print(note_list_x)
    # print(note_list_y)
    # print(note_list)

    note_submit = [i for i, x in enumerate(note_submit)if x == None]


    # note_submit_1 = [i - 1 for i in note_submit]
    print(note_submit)
    print("kore")



        

    newcnt = 0
    until_submit_volume = len(until_submit)
    for i in range(until_submit_volume):
        A = until_submit_1[newcnt]
        # print(A)
        B = newkadai[A]
        st.sidebar.write(B)
        newcnt = newcnt + 1
    

    newcnt = 0
    note_submit_volume = len(note_submit)
    # print(note_submit)
    
    print(note_list)
    print(note_submit_volume)
    
    for i in range(note_submit_volume):
        C = note_submit[newcnt]
        D = note_list[C]

        st.sidebar.write(D)
        newcnt = newcnt + 1

    
    if len(until_submit) < 1 and note_submit_volume < 1:
        st.sidebar.write("すべて提出されています!")


    # print(note_submit)   


print("fin")

book.close()

book.save('sample.xlsx')

